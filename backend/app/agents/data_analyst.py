"""
Data Analyst Agent â€” Analyzes structured data uploaded alongside RFPs.

Processes ticket dumps (CSV/Excel), volume reports, SLA dashboards, and any
tabular data to extract operational insights that feed into downstream agents.

Output:
- Application-wise ticket volume analysis
- Priority/severity distribution
- Trend analysis (monthly volumes, resolution times)
- Complexity indicators for scope and staffing
- Automation opportunity signals

Works for any contract type and any number of applications.
"""

import os
from typing import Any, Dict, List, Optional
from app.agents.base import BaseAgent


class DataAnalystAgent(BaseAgent):
    name = "Data Analyst Agent"
    agent_tier = "analytical"

    @staticmethod
    def _safe_int(value) -> int:
        """Coerce a value to int, treating None/non-numeric as 0."""
        if value is None:
            return 0
        try:
            return int(value)
        except (TypeError, ValueError):
            return 0

    def __init__(
        self, bid_id: str, manifest: Dict[str, Any], document_texts: List[str] = None
    ):
        super().__init__(bid_id, manifest)
        self.document_texts = document_texts or []
        self.structured_data = []  # Parsed tabular data

    async def observe(self) -> Dict[str, Any]:
        """Detect and classify all uploaded data â€” separate RFP narrative from structured data."""
        structured_texts = []
        narrative_texts = []

        for text in self.document_texts:
            if self._is_structured_data(text):
                structured_texts.append(text)
            else:
                narrative_texts.append(text)

        # Also check for structured data from file paths in manifest
        docs = self.manifest.get("documents", [])
        for doc in docs:
            path = doc.get("file_path", "")
            ext = os.path.splitext(path)[1].lower()
            if ext in (".csv", ".xlsx", ".xls"):
                try:
                    parsed = self._parse_structured_file(path)
                    if parsed:
                        structured_texts.append(parsed["text"])
                        self.structured_data.append(parsed)
                except Exception as e:
                    self.log("file_parse_error", {"path": path, "error": str(e)})

        self.log(
            "data_classified",
            {
                "structured_docs": len(structured_texts),
                "narrative_docs": len(narrative_texts),
            },
        )

        return {
            "structured_texts": structured_texts,
            "narrative_texts": narrative_texts,
            "has_structured_data": len(structured_texts) > 0,
            "intake": self.manifest.get("intake_output", {}),
        }

    async def orient(self, obs: Dict[str, Any]) -> Dict[str, Any]:
        """Analyze structured data to extract operational intelligence."""
        if not obs["has_structured_data"]:
            # No structured data â€” generate minimal analysis from RFP narrative
            return await self._analyze_from_narrative(obs)

        # Combine structured data for analysis
        combined_data = "\n\n".join(obs["structured_texts"][:50000])

        # Step 1: Detect data structure (columns, types, patterns)
        detect_prompt = f"""Analyze this structured data dump. Identify what it contains.

DATA (first 8000 chars):
{combined_data[:8000]}

Return JSON:
{{
  "data_type": "ticket_dump|volume_report|sla_dashboard|cr_log|incident_log|resource_plan|other",
  "detected_columns": ["column names or inferred column purposes"],
  "total_rows_estimated": 0,
  "applications_found": ["application/system names found in data"],
  "date_range": {{"start": "earliest date", "end": "latest date"}},
  "key_patterns": ["notable patterns in the data"],
  "data_quality": "high|medium|low"
}}"""

        structure = await self.llm_json(detect_prompt, max_tokens=3000)
        self.log(
            "structure_detected",
            {
                "type": structure.get("data_type"),
                "apps": len(structure.get("applications_found", [])),
                "rows": structure.get("total_rows_estimated"),
            },
        )

        # Step 2: Deep analysis based on data type
        analysis_prompt = f"""You are an expert operations analyst. Analyze this data and extract actionable intelligence for a bid response.

DATA TYPE: {structure.get("data_type", "unknown")}
DETECTED APPS: {", ".join(structure.get("applications_found", []))}
COLUMNS: {", ".join(structure.get("detected_columns", []))}

FULL DATA (up to 20000 chars):
{combined_data[:20000]}

Return JSON with ALL of these sections:
{{
  "application_analysis": [
    {{
      "application": "app name",
      "ticket_count": 0,
      "percentage_of_total": 0,
      "priority_distribution": {{"critical": 0, "high": 0, "medium": 0, "low": 0}},
      "top_categories": ["category1", "category2"],
      "avg_resolution_hours": 0,
      "complexity_rating": "high|medium|low",
      "automation_potential": "high|medium|low",
      "key_insights": ["insight1"]
    }}
  ],
  "volume_summary": {{
    "total_tickets": 0,
    "monthly_average": 0,
    "peak_month": "month name",
    "peak_volume": 0,
    "trend": "increasing|stable|decreasing",
    "by_priority": {{"critical": 0, "high": 0, "medium": 0, "low": 0}},
    "by_type": {{"incident": 0, "service_request": 0, "change_request": 0, "problem": 0, "enhancement": 0}}
  }},
  "staffing_indicators": {{
    "estimated_fte_needed": 0,
    "rationale": "how estimated",
    "peak_staffing": 0,
    "recommended_shift_model": "description"
  }},
  "automation_opportunities": [
    {{
      "area": "what to automate",
      "ticket_volume_impact": 0,
      "percentage_reduction": 0,
      "complexity": "low|medium|high",
      "justification": "why"
    }}
  ],
  "risk_indicators": [
    {{
      "risk": "description",
      "evidence": "from data",
      "impact": "high|medium|low",
      "recommendation": "action"
    }}
  ],
  "sla_performance_baseline": {{
    "current_resolution_adherence": "percentage or description",
    "response_time_avg": "value",
    "mttr": "mean time to resolve",
    "recurring_issues": ["issue1"]
  }}
}}

Be specific with numbers. Use actual counts from the data. If data is insufficient for a field, set to null."""

        analysis = await self.llm_json(analysis_prompt, max_tokens=4000)
        self.log(
            "analysis_complete",
            {
                "apps_analyzed": len(analysis.get("application_analysis", [])),
                "total_tickets": analysis.get("volume_summary", {}).get(
                    "total_tickets", 0
                ),
            },
        )

        return {
            "structure": structure,
            "analysis": analysis,
            "has_data": True,
        }

    async def _analyze_from_narrative(self, obs: Dict[str, Any]) -> Dict[str, Any]:
        """Extract ALL quantitative and operational data from the RFP document."""
        # Prefer indexed sections; fall back to raw text only if index is empty
        rfp_text = self.get_rfp_sections("data_analyst", max_chars=15000)
        if not rfp_text or len(rfp_text) < 200:
            rfp_text = self.manifest.get("rfp_text", "")
        if not rfp_text:
            return {"has_data": False, "analysis": {}}

        # Cap context to prevent token bloat
        text_chunk = rfp_text[:15000]

        prompt = f"""You are a senior data analyst preparing a comprehensive data intelligence report for a bid response team.

Analyze every piece of quantitative and operational data in this RFP document. Extract ALL numbers, metrics, counts, and factual data points.

RFP DOCUMENT:
{text_chunk}

Return a comprehensive JSON with ALL of these sections. Fill every field with actual data from the document. If data is not explicitly stated, infer from context and mark confidence as "inferred":

{{
  "user_landscape": {{
    "total_users": 0,
    "users_by_country": [{{"country": "name", "user_count": 0, "notes": ""}}],
    "total_countries": 0,
    "regions": ["region names"],
    "user_growth_trend": "description"
  }},
  "application_landscape": {{
    "platforms": [
      {{
        "name": "platform name",
        "vendor": "vendor",
        "purpose": "what it does",
        "user_count": 0,
        "country_count": 0,
        "modules": ["module names"],
        "version": "if mentioned",
        "hosting": "cloud/on-prem/hybrid"
      }}
    ],
    "total_applications": 0,
    "integration_count": 0,
    "integrations": [
      {{"source": "system A", "target": "system B", "type": "API/file/middleware", "frequency": "real-time/batch", "criticality": "high/medium/low"}}
    ]
  }},
  "configuration_complexity": {{
    "pay_rules": 0,
    "work_rules": 0,
    "business_rules": 0,
    "custom_configurations": 0,
    "custom_reports": 0,
    "workflows": 0,
    "templates": 0,
    "total_config_items": 0,
    "complexity_rating": "high/medium/low",
    "notes": ["relevant config details"]
  }},
  "operational_volumes": {{
    "incidents_per_month": 0,
    "change_requests_per_month": 0,
    "enhancement_requests_per_year": 0,
    "releases_per_year": 0,
    "total_tickets_annual": 0,
    "sla_targets": [
      {{"metric": "name", "target": "value", "penalty": "if any"}}
    ],
    "uptime_requirement": "percentage",
    "support_hours": "description",
    "response_time_targets": {{"critical": "", "high": "", "medium": "", "low": ""}}
  }},
  "contract_dimensions": {{
    "contract_duration_months": 0,
    "contract_type": "AMS/Implementation/Advisory",
    "estimated_value_range": "if mentioned",
    "pricing_model": "FP/T&M/hybrid",
    "payment_terms": "description",
    "renewal_options": "description"
  }},
  "current_state": {{
    "current_provider": "if mentioned",
    "current_team_size": 0,
    "current_team_structure": "description",
    "pain_points": ["identified issues"],
    "transition_requirements": "description",
    "knowledge_transfer_period": "duration"
  }},
  "key_metrics_summary": [
    {{"metric": "descriptive name", "value": "number or text", "source": "where found in RFP", "confidence": "explicit/inferred"}}
  ],
  "data_quality_assessment": {{
    "completeness": "high/medium/low",
    "data_gaps": ["what's missing"],
    "recommendations": ["what additional data to request"]
  }}
}}

CRITICAL: Extract EVERY number mentioned in the RFP. Include user counts, country counts, configuration counts, integration counts, SLA targets, team sizes, timelines, volumes â€” everything quantifiable. Be thorough and precise."""

        analysis = await self.llm_json(prompt, max_tokens=5000)

        # Build a rich structure response
        platforms = analysis.get("application_landscape", {}).get("platforms", [])
        structure = {
            "data_type": "rfp_comprehensive_analysis",
            "detected_columns": list(analysis.keys()),
            "total_rows_estimated": len(analysis.get("key_metrics_summary", [])),
            "applications_found": [p.get("name", "") for p in platforms],
            "data_quality": analysis.get("data_quality_assessment", {}).get(
                "completeness", "medium"
            ),
        }

        return {"has_data": True, "analysis": analysis, "structure": structure}

    async def decide(self, orientation: Dict[str, Any]) -> Dict[str, Any]:
        return orientation

    async def act(self, decision: Dict[str, Any]) -> Dict[str, Any]:
        analysis = decision.get("analysis", {})
        structure = decision.get("structure", {})
        has_data = decision.get("has_data", False)

        if not has_data:
            return {
                "data_analysis": {},
                "narrative": "No operational data was available for analysis.",
                "hitl_summary": "Data Intelligence: No data available. Upload RFP or structured data files.",
            }

        # Extract key metrics from the comprehensive analysis
        user_data = analysis.get("user_landscape", {})
        app_data = analysis.get("application_landscape", {})
        config_data = analysis.get("configuration_complexity", {})
        ops_data = analysis.get("operational_volumes", {})
        contract_data = analysis.get("contract_dimensions", {})
        current_state = analysis.get("current_state", {})

        total_users = self._safe_int(user_data.get("total_users", 0))
        total_countries = self._safe_int(user_data.get("total_countries", 0))
        platforms = app_data.get("platforms", []) or []
        integrations = app_data.get("integrations", []) or []
        total_configs = self._safe_int(config_data.get("total_config_items", 0))
        sla_targets = ops_data.get("sla_targets", []) or []
        user_data.get("users_by_country", []) or []
        key_metrics = analysis.get("key_metrics_summary", []) or []

        # === INTAKE FALLBACK WIRING ===
        # If LLM extraction returned empty/0 for key metrics, pull from intake
        # which already parsed the RFP with high fidelity.
        intake = self.manifest.get("intake_output", {})
        if intake and isinstance(intake, dict):
            ef = intake.get("extracted_fields", intake)
            intake_platforms = intake.get("platform_details", [])
            intake_integrations = intake.get("integration_inventory", [])
            intake_slas = intake.get("kpi_sla_table", [])

            # Fill platforms from intake if LLM returned empty
            if not platforms and intake_platforms:
                platforms = [
                    {
                        "name": p.get("product_name", "Unknown"),
                        "vendor": p.get("vendor", ""),
                        "purpose": p.get("description", ""),
                        "user_count": p.get("user_count", 0),
                        "country_count": len(p.get("countries_deployed", []))
                        if isinstance(p.get("countries_deployed"), list)
                        else 0,
                        "modules": p.get("modules", []),
                    }
                    for p in intake_platforms
                    if isinstance(p, dict)
                ]
                app_data["platforms"] = platforms

            # Fill integrations from intake if empty
            if not integrations and intake_integrations:
                integrations = [
                    {
                        "source": i.get("source", ""),
                        "target": i.get("target", ""),
                        "type": i.get("type", ""),
                        "frequency": i.get("frequency", ""),
                        "criticality": i.get("criticality", "medium"),
                    }
                    for i in intake_integrations
                    if isinstance(i, dict)
                ]
                app_data["integrations"] = integrations

            # Fill SLA targets from intake if empty
            if not sla_targets and intake_slas:
                sla_targets = [
                    {
                        "metric": s.get("metric", s.get("kpi", "")),
                        "target": s.get("target", s.get("value", "")),
                        "penalty": s.get("penalty", ""),
                    }
                    for s in intake_slas
                    if isinstance(s, dict)
                ]
                ops_data["sla_targets"] = sla_targets

            # Fill user/country counts from intake if 0
            if isinstance(ef, dict):
                if total_users == 0:
                    emp = ef.get("employee_population", {})
                    if isinstance(emp, dict):
                        total_users = self._safe_int(emp.get("value", 0))
                    elif isinstance(emp, (int, float)):
                        total_users = self._safe_int(emp)
                    else:
                        total_users = self._safe_int(emp)
                    user_data["total_users"] = total_users

                if total_countries == 0:
                    geo = ef.get("geographies", ef.get("client_geography", {}))
                    if isinstance(geo, dict):
                        geo_val = geo.get("value", [])
                    elif isinstance(geo, list):
                        geo_val = geo
                    else:
                        geo_val = []
                    total_countries = len(geo_val) if isinstance(geo_val, list) else 0
                    user_data["total_countries"] = total_countries

            # Fill config items from intake if 0
            if total_configs == 0 and intake_platforms:
                # Estimate config complexity from platform module counts
                total_configs = sum(
                    len(p.get("modules", []))
                    for p in intake_platforms
                    if isinstance(p, dict)
                )
                config_data["total_config_items"] = total_configs

            self.log(
                "intake_fallback_applied",
                {
                    "platforms": len(platforms),
                    "integrations": len(integrations),
                    "sla_targets": len(sla_targets),
                    "users": total_users,
                    "countries": total_countries,
                },
            )

        # Generate executive narrative
        prompt = f"""Write a data intelligence brief for a bid response team.

STYLE: No markdown. No bullets. No asterisks. Professional flowing paragraphs with embedded numbers. Client-ready.

DATA:
- Users: {total_users} across {total_countries} countries
- Platforms: {", ".join([p.get("name", "?") for p in platforms[:8]])}
- Integrations: {len(integrations)} | Config Items: {total_configs}
- SLA Targets: {len(sla_targets)} | Key Metrics: {len(key_metrics)}
- Contract: {contract_data.get("contract_duration_months", "unknown")} months
- Provider: {current_state.get("current_provider", "unknown")}

Write exactly 3 paragraphs (3-4 sentences each). Complete every sentence:
1. Scale and landscape â€” users, countries, platforms, what was extracted.
2. Integration and configuration complexity â€” system interdependencies, rule volumes.
3. Strategic implications â€” what this means for staffing, pricing, risk.

Do NOT stop mid-sentence. Complete all 3 paragraphs."""

        narrative = await self.llm_generate(prompt, max_tokens=1500)

        # Build HITL summary with real numbers
        platform_names = ", ".join([p.get("name", "?") for p in platforms[:5]])
        summary = f"Data Intelligence: {len(platforms)} platforms ({platform_names}). "
        summary += f"{total_users:,} users across {total_countries} countries. "
        summary += f"{len(integrations)} integrations, {total_configs} config items. "
        summary += f"{len(sla_targets)} SLA targets. "
        summary += f"{len(key_metrics)} key metrics extracted."

        # === CLOSED-LOOP: Capture learnings for institutional improvement ===
        if platforms:
            self.capture_learning(
                learning_type="data_pattern",
                insight=f"RFP with {len(platforms)} platforms, {len(integrations)} integrations, "
                f"{total_users} users across {total_countries} countries. "
                f"Config complexity: {total_configs} items. "
                f"Data quality: {structure.get('data_quality', 'unknown')}.",
                confidence=0.7,
            )
        if not decision.get("has_data") or (total_users == 0 and not platforms):
            self.capture_learning(
                learning_type="process_gap",
                insight="Data Analyst received no structured data or RFP narrative. "
                "Ensure RFP documents are uploaded before pipeline execution.",
                confidence=0.6,
            )

        return {
            "data_analysis": analysis,
            "data_structure": structure,
            "narrative": narrative,
            "hitl_summary": summary,
        }

    @staticmethod
    def _is_structured_data(text: str) -> bool:
        """Detect if text looks like structured/tabular data vs narrative."""
        lines = text.strip().split("\n")[:20]
        if not lines:
            return False

        # Check for CSV-like patterns (pipe-separated or comma-separated)
        pipe_lines = sum(1 for l in lines if l.count("|") >= 3)
        comma_lines = sum(1 for l in lines if l.count(",") >= 3)
        tab_lines = sum(1 for l in lines if l.count("\t") >= 2)

        # If >50% of sampled lines look tabular, it's structured
        threshold = len(lines) * 0.5
        if pipe_lines > threshold or comma_lines > threshold or tab_lines > threshold:
            return True

        # Check for ticket-dump patterns
        ticket_keywords = [
            "ticket",
            "incident",
            "priority",
            "severity",
            "status",
            "resolution",
            "assigned",
            "created",
            "closed",
            "category",
        ]
        first_line_lower = lines[0].lower() if lines else ""
        keyword_hits = sum(1 for kw in ticket_keywords if kw in first_line_lower)
        if keyword_hits >= 3:
            return True

        return False

    @staticmethod
    def _parse_structured_file(file_path: str) -> Optional[Dict[str, Any]]:
        """Parse CSV/Excel files into structured format."""
        ext = os.path.splitext(file_path)[1].lower()

        if ext == ".csv":
            try:
                with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                    text = f.read()
                lines = text.strip().split("\n")
                return {
                    "text": text,
                    "format": "csv",
                    "row_count": len(lines) - 1,
                    "header": lines[0] if lines else "",
                }
            except Exception:
                return None

        elif ext in (".xlsx", ".xls"):
            try:
                from openpyxl import load_workbook

                wb = load_workbook(file_path, data_only=True)
                all_text = []
                total_rows = 0
                for sheet in wb.worksheets:
                    for row in sheet.iter_rows(values_only=True):
                        row_text = " | ".join(str(c) for c in row if c is not None)
                        if row_text.strip():
                            all_text.append(row_text)
                            total_rows += 1
                return {
                    "text": "\n".join(all_text),
                    "format": "xlsx",
                    "row_count": total_rows,
                    "header": all_text[0] if all_text else "",
                }
            except Exception:
                return None

        return None
