"""
Excel Generator — generates comprehensive commercial models and bid summary workbooks.
Dynamically extracts data from ALL agent outputs rather than hardcoding sections.
"""

import os
import logging
from typing import Any, Dict, Optional
from datetime import datetime


def _unwrap(output) -> dict:
    """Unwrap agent output envelope {status, agent, result: {...}} if needed."""
    if not output or not isinstance(output, dict):
        return {}
    if "result" in output and isinstance(output.get("result"), dict):
        if "status" in output or "agent" in output:
            return output["result"]
    return output


# Maps DB column names to display-friendly labels
AGENT_DISPLAY_NAMES = {
    "intake_output": "Intake Analysis",
    "data_analyst_output": "Data Analyst",
    "client_intel_output": "Client Intelligence",
    "bid_no_bid_output": "Strategic Assessment",
    "scope_output": "Scope Builder",
    "solution_output": "Solution Architect",
    "automation_ai_output": "Automation & AI",
    "competitive_output": "Competitive Intelligence",
    "commercial_output": "Commercial Model",
    "compliance_output": "Compliance & Risk",
    "proposal_output": "Proposal Writer",
    "output_generator_output": "Output Generator",
    "discovery_output": "Discovery & Clarifications",
    "qa_output": "Quality Assurance",
    "feedback_output": "Feedback & Learning",
    "transition_change_output": "Transition & Change",
}


def _flatten_dict(
    d: dict, prefix: str = "", max_depth: int = 3, depth: int = 0
) -> list:
    """Flatten a nested dict into [(key_path, value)] pairs for table rendering."""
    rows = []
    if depth >= max_depth or not isinstance(d, dict):
        return rows
    for k, v in d.items():
        key_path = f"{prefix}.{k}" if prefix else k
        if isinstance(v, dict) and depth < max_depth - 1:
            rows.extend(_flatten_dict(v, key_path, max_depth, depth + 1))
        elif isinstance(v, list):
            if v and isinstance(v[0], dict):
                rows.append((key_path, f"[{len(v)} items]"))
            else:
                rows.append((key_path, ", ".join(str(x) for x in v[:10])))
        else:
            rows.append((key_path, v))
    return rows


def _write_flat_data(
    ws, data: dict, header_font, header_fill, thin_border, start_row: int = 1
) -> int:
    """Write a flat key-value section into a worksheet. Returns last row used."""
    rows = _flatten_dict(data)
    for i, (key, value) in enumerate(rows):
        row_num = start_row + i
        ws.cell(
            row=row_num, column=1, value=key.replace("_", " ").title()
        ).border = thin_border
        ws.cell(row=row_num, column=1).font = header_font
        cell = ws.cell(
            row=row_num, column=2, value=str(value) if value is not None else ""
        )
        cell.border = thin_border
    return start_row + len(rows)


def _write_list_table(
    ws,
    items: list,
    sheet_title: str,
    header_font_white,
    header_fill,
    thin_border,
    start_row: int = 1,
) -> int:
    """Write a list of dicts as a table. Returns last row used."""
    if not items or not isinstance(items[0], dict):
        return start_row

    headers = list(items[0].keys())
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=h.replace("_", " ").title())
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, item in enumerate(items[:50], start_row + 1):  # Max 50 rows
        for col_idx, h in enumerate(headers, 1):
            val = item.get(h, "")
            if isinstance(val, (dict, list)):
                val = str(val)[:200]
            cell = ws.cell(
                row=row_idx, column=col_idx, value=str(val) if val is not None else ""
            )
            cell.border = thin_border

    return start_row + len(items) + 1


def generate_commercial_model(
    bid_data: Dict[str, Any],
    output_dir: str = "../knowledge_base",
    filename: Optional[str] = None,
) -> Dict[str, Any]:
    """Generate a comprehensive commercial model Excel workbook from all agent outputs."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    except ImportError:
        return {"status": "error", "message": "openpyxl not installed"}

    wb = Workbook()
    manifest = bid_data.get("manifest", {})
    client = manifest.get("client", {})
    client_name = client.get("name") or bid_data.get("client_name", "Client")

    header_font = Font(bold=True, size=11)
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(
        start_color="0066FF", end_color="0066FF", fill_type="solid"
    )
    section_fill = PatternFill(
        start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"
    )
    green_fill = PatternFill(
        start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"
    )
    amber_fill = PatternFill(
        start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"
    )
    red_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ══════════════════════════════════════════════════════
    # SHEET 1: EXECUTIVE SUMMARY
    # ══════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Executive Summary"
    ws["A1"] = f"Bid Response — {client_name}"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")

    bid_ref = bid_data.get("bid_reference", "")
    industry = bid_data.get("client_industry", "")
    wp = bid_data.get("win_probability")
    rec = bid_data.get("bid_recommendation", "")
    tcv = bid_data.get("estimated_tcv")

    summary_data = [
        ("", ""),
        ("BID OVERVIEW", ""),
        ("Bid Reference", bid_ref),
        ("Client Name", client_name),
        ("Industry", industry),
        (
            "Win Probability",
            f"{wp * 100:.0f}%" if isinstance(wp, (int, float)) and wp else "N/A",
        ),
        ("Recommendation", rec or "N/A"),
        (
            "Estimated TCV",
            f"${tcv:,.0f}" if isinstance(tcv, (int, float)) and tcv else "N/A",
        ),
        ("Status", bid_data.get("status", "")),
    ]

    for i, (label, val) in enumerate(summary_data, start=4):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = val
        if label and label == label.upper() and val == "":
            ws[f"A{i}"].font = Font(bold=True, size=12, color="0066FF")
            ws[f"A{i}"].fill = section_fill
            ws[f"B{i}"].fill = section_fill
        elif label:
            ws[f"A{i}"].font = header_font
            ws[f"A{i}"].border = thin_border
            ws[f"B{i}"].border = thin_border

    # ══════════════════════════════════════════════════════
    # SHEET 2: COMMERCIAL MODEL (dedicated deep-dive)
    # ══════════════════════════════════════════════════════
    commercial = _unwrap(bid_data.get("commercial_output"))
    ws2 = wb.create_sheet("Commercial Model")
    ws2["A1"] = f"Commercial Model — {client_name}"
    ws2["A1"].font = Font(bold=True, size=14)

    pl = commercial.get("pl_model", {})
    revenue = pl.get("revenue", {})
    costs = pl.get("costs", {})
    profit = pl.get("profitability", {})
    params = commercial.get("contract_params", {})

    comm_data = [
        ("", ""),
        ("CONTRACT PARAMETERS", ""),
        ("Contract Duration (months)", params.get("months", 0)),
        ("Pricing Model", params.get("pricing_model", "N/A")),
        ("Currency", params.get("currency", "USD")),
        ("", ""),
        ("REVENUE", ""),
        ("Total Contract Value", revenue.get("total_contract_value", 0)),
        ("Annual Revenue", revenue.get("annual_revenue", 0)),
        ("Monthly Price", revenue.get("monthly_price", 0)),
        ("", ""),
        ("COSTS", ""),
        ("Direct Delivery", costs.get("direct_delivery", 0)),
        ("Transition", costs.get("transition", 0)),
        ("Tools & Infrastructure", costs.get("tools_infra", 0)),
        ("Travel", costs.get("travel", 0)),
        ("Contingency", costs.get("contingency", 0)),
        ("SG&A", costs.get("sga", 0)),
        ("Total COGS", costs.get("total_cogs", 0)),
        ("", ""),
        ("PROFITABILITY", ""),
        ("Gross Profit", profit.get("gross_profit", 0)),
        (
            "Margin %",
            f"{profit.get('margin_percent', 0):.1f}%"
            if isinstance(profit.get("margin_percent"), (int, float))
            else "N/A",
        ),
        ("Target Margin", f"{profit.get('target_margin', 25)}%"),
    ]
    for i, (label, val) in enumerate(comm_data, start=3):
        ws2[f"A{i}"] = label
        ws2[f"B{i}"] = val
        if label and label == label.upper() and val == "":
            ws2[f"A{i}"].font = Font(bold=True, size=11, color="0066FF")
            ws2[f"A{i}"].fill = section_fill
            ws2[f"B{i}"].fill = section_fill
        elif label:
            ws2[f"A{i}"].font = header_font
            ws2[f"A{i}"].border = thin_border
            ws2[f"B{i}"].border = thin_border

    # Margin guardrail indicator
    guardrail = commercial.get("margin_guardrail", {})
    row = len(comm_data) + 5
    if guardrail:
        ws2[f"A{row}"] = "MARGIN GUARDRAIL"
        ws2[f"A{row}"].font = Font(bold=True, size=11)
        ws2[f"A{row + 1}"] = guardrail.get("message", "")
        status = guardrail.get("status", "pass")
        fill = (
            green_fill
            if status == "pass"
            else amber_fill
            if status == "flag"
            else red_fill
        )
        ws2[f"A{row + 1}"].fill = fill
        ws2[f"B{row + 1}"].fill = fill

    # ══════════════════════════════════════════════════════
    # SHEET 3: RESOURCE PLAN
    # ══════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Resource Plan")
    res_headers = ["Role", "Location", "FTE Count", "Start Month", "Justification"]
    for col, h in enumerate(res_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    resources = commercial.get("resource_plan", [])
    for row_idx, r in enumerate(resources, 2):
        if isinstance(r, dict):
            ws3.cell(
                row=row_idx, column=1, value=r.get("role", "")
            ).border = thin_border
            ws3.cell(
                row=row_idx, column=2, value=r.get("location", "")
            ).border = thin_border
            ws3.cell(
                row=row_idx, column=3, value=r.get("count", 0)
            ).border = thin_border
            ws3.cell(
                row=row_idx, column=4, value=r.get("start_month", 1)
            ).border = thin_border
            ws3.cell(
                row=row_idx, column=5, value=r.get("justification", "")
            ).border = thin_border

    total_row = len(resources) + 2
    ws3.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws3.cell(
        row=total_row,
        column=3,
        value=sum(r.get("count", 0) for r in resources if isinstance(r, dict)),
    ).font = Font(bold=True)

    # ══════════════════════════════════════════════════════
    # SHEET 4: SCENARIO COMPARISON
    # ══════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Scenario Comparison")
    scenarios = commercial.get("scenarios", {})

    sc_headers = ["Metric", "Base Case", "Aggressive", "Conservative"]
    for col, h in enumerate(sc_headers, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    scenario_rows = [
        (
            "Total Contract Value",
            scenarios.get("base", {}).get("tcv", 0),
            scenarios.get("aggressive", {}).get("tcv", 0),
            scenarios.get("conservative", {}).get("tcv", 0),
        ),
        (
            "Margin %",
            scenarios.get("base", {}).get("margin", 0),
            scenarios.get("aggressive", {}).get("margin", 0),
            scenarios.get("conservative", {}).get("margin", 0),
        ),
        ("Target Margin", "25%", "18%", "30%"),
        ("Resource Adjustment", "Standard", "-10% FTE", "+15% FTE"),
        ("Contingency", "5%", "3%", "8%"),
    ]
    for row_idx, row_data in enumerate(scenario_rows, 2):
        for col_idx, val in enumerate(row_data):
            cell = ws4.cell(row=row_idx, column=col_idx + 1, value=val)
            cell.border = thin_border
            if col_idx == 0:
                cell.font = Font(bold=True)

    # ══════════════════════════════════════════════════════
    # SHEET 5: SCOPE BREAKDOWN
    # ══════════════════════════════════════════════════════
    scope = _unwrap(bid_data.get("scope_output"))
    scope_pkg = scope.get("scope_package", {}) if isinstance(scope, dict) else {}
    scope_by_platform = scope_pkg.get("scope_by_platform", [])

    if scope_by_platform:
        ws5 = wb.create_sheet("Scope Breakdown")
        sp_headers = ["Platform", "Effort (Days)", "FTE", "Scope Summary"]
        for col, h in enumerate(sp_headers, 1):
            cell = ws5.cell(row=1, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
        for row_idx, plat in enumerate(scope_by_platform, 2):
            if isinstance(plat, dict):
                ws5.cell(
                    row=row_idx, column=1, value=plat.get("platform", "")
                ).border = thin_border
                ws5.cell(
                    row=row_idx, column=2, value=plat.get("platform_effort_days", 0)
                ).border = thin_border
                ws5.cell(
                    row=row_idx, column=3, value=plat.get("fte", "")
                ).border = thin_border
                ws5.cell(
                    row=row_idx,
                    column=4,
                    value=str(plat.get("scope_summary", ""))[:200],
                ).border = thin_border

    # ══════════════════════════════════════════════════════
    # SHEET 6: RISK REGISTER
    # ══════════════════════════════════════════════════════
    compliance = _unwrap(bid_data.get("compliance_output"))
    risks = compliance.get("contractual_risks", compliance.get("risk_register", []))
    if risks and isinstance(risks, list):
        ws6 = wb.create_sheet("Risk Register")
        risk_headers = ["ID", "Risk", "Severity", "Likelihood", "Mitigation"]
        for col, h in enumerate(risk_headers, 1):
            cell = ws6.cell(row=1, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
        for row_idx, risk in enumerate(risks, 2):
            if isinstance(risk, dict):
                ws6.cell(
                    row=row_idx, column=1, value=risk.get("id", risk.get("clause", ""))
                ).border = thin_border
                ws6.cell(
                    row=row_idx,
                    column=2,
                    value=str(risk.get("description", risk.get("risk", "")))[:200],
                ).border = thin_border
                ws6.cell(
                    row=row_idx,
                    column=3,
                    value=risk.get("severity", risk.get("impact", "")),
                ).border = thin_border
                ws6.cell(
                    row=row_idx, column=4, value=risk.get("likelihood", "")
                ).border = thin_border
                ws6.cell(
                    row=row_idx, column=5, value=str(risk.get("mitigation", ""))[:200]
                ).border = thin_border

    # ══════════════════════════════════════════════════════
    # SHEET 7: AUTOMATION OPPORTUNITIES
    # ══════════════════════════════════════════════════════
    automation = _unwrap(bid_data.get("automation_ai_output"))
    opportunities = automation.get("opportunities", [])
    if opportunities and isinstance(opportunities, list):
        ws7 = wb.create_sheet("Automation Opportunities")
        auto_headers = ["Title", "Platform", "Priority", "Savings", "Description"]
        for col, h in enumerate(auto_headers, 1):
            cell = ws7.cell(row=1, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
        for row_idx, opp in enumerate(opportunities, 2):
            if isinstance(opp, dict):
                ws7.cell(
                    row=row_idx, column=1, value=opp.get("title", opp.get("name", ""))
                ).border = thin_border
                ws7.cell(
                    row=row_idx, column=2, value=opp.get("platform", "")
                ).border = thin_border
                ws7.cell(
                    row=row_idx, column=3, value=opp.get("priority", "")
                ).border = thin_border
                ws7.cell(
                    row=row_idx,
                    column=4,
                    value=str(
                        opp.get("estimated_savings", opp.get("effort_reduction", ""))
                    ),
                ).border = thin_border
                ws7.cell(
                    row=row_idx, column=5, value=str(opp.get("description", ""))[:200]
                ).border = thin_border

    # ══════════════════════════════════════════════════════
    # SHEET 8+: DYNAMIC AGENT OUTPUT DUMP (for any agent not already covered)
    # ══════════════════════════════════════════════════════
    covered_agents = {
        "commercial_output",
        "scope_output",
        "compliance_output",
        "automation_ai_output",
    }

    for agent_key, agent_label in AGENT_DISPLAY_NAMES.items():
        if agent_key in covered_agents:
            continue

        raw = bid_data.get(agent_key)
        if not raw:
            continue

        data = _unwrap(raw)
        if not data:
            continue

        # Create a sheet for this agent's output
        sheet_name = agent_label[:31]  # Excel max sheet name length
        try:
            ws_agent = wb.create_sheet(sheet_name)
            ws_agent["A1"] = f"{agent_label} Output"
            ws_agent["A1"].font = Font(bold=True, size=14)

            # Check for common list-based structures to render as tables
            rendered_table = False
            for list_key in data:
                val = data[list_key]
                if isinstance(val, list) and val and isinstance(val[0], dict):
                    _write_list_table(
                        ws_agent,
                        val,
                        list_key,
                        header_font_white,
                        header_fill,
                        thin_border,
                        start_row=3,
                    )
                    rendered_table = True
                    break

            if not rendered_table:
                _write_flat_data(
                    ws_agent, data, header_font, header_fill, thin_border, start_row=3
                )
        except Exception as e:
            logging.warning(f"[EXCEL] Failed to create sheet for {agent_key}: {e}")

    # ══════════════════════════════════════════════════════
    # QA SCORES SHEET
    # ══════════════════════════════════════════════════════
    qa = _unwrap(bid_data.get("qa_output"))
    qa_scores = qa.get("quality_scores", {})
    if qa_scores:
        ws_qa = wb.create_sheet("QA Scores")
        ws_qa["A1"] = "Quality Assurance Scores"
        ws_qa["A1"].font = Font(bold=True, size=14)
        qa_headers = ["Dimension", "Score", "Max"]
        for col, h in enumerate(qa_headers, 1):
            cell = ws_qa.cell(row=3, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
        for row_idx, (dim, score) in enumerate(qa_scores.items(), 4):
            ws_qa.cell(
                row=row_idx, column=1, value=dim.replace("_", " ").title()
            ).border = thin_border
            ws_qa.cell(row=row_idx, column=2, value=score).border = thin_border
            ws_qa.cell(row=row_idx, column=3, value=100).border = thin_border

    # Auto-width all columns across all sheets
    for ws_item in wb.worksheets:
        try:
            for col in ws_item.columns:
                col_cells = list(col)
                if col_cells:
                    max_len = max(len(str(cell.value or "")) for cell in col_cells)
                    ws_item.column_dimensions[col_cells[0].column_letter].width = min(
                        max_len + 4, 50
                    )
        except Exception:
            pass

    # Save
    os.makedirs(output_dir, exist_ok=True)
    fname = (
        filename
        or f"Commercial_{client_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )
    path = os.path.join(output_dir, fname)
    wb.save(path)

    return {
        "status": "success",
        "path": path,
        "filename": fname,
        "sheets": len(wb.sheetnames),
    }
